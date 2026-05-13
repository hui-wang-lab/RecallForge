"""CSV/TSV/XLSX table file parser."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from recallforge.chunking.core.ids import block_id
from recallforge.chunking.ir.models import Block, Page, ParsedDocument, ParseReport
from recallforge.chunking.parsers.base import ParserAdapter, ParserConfig
from recallforge.chunking.parsers.utils import file_document_id


class TableFileParser(ParserAdapter):
    name = "table_file"

    def is_available(self) -> bool:
        return True

    def parse(self, path: str | Path, config: ParserConfig) -> ParsedDocument:
        path_obj = Path(path)
        suffix = path_obj.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            rows_by_sheet = {path_obj.stem: _read_delimited(path_obj, delimiter="\t" if suffix == ".tsv" else None)}
        elif suffix in {".xlsx", ".xlsm"}:
            rows_by_sheet = _read_xlsx(path_obj)
        else:
            raise RuntimeError(f"Unsupported table file type: {suffix}")

        document_id = file_document_id(path_obj)
        blocks: list[Block] = []
        reading_order = 0

        for sheet_name, rows in rows_by_sheet.items():
            if not rows:
                continue
            headers = _headers(rows[0])
            for row_offset, row in enumerate(rows[1:], start=2):
                values = _row_dict(headers, row)
                if not any(value for value in values.values()):
                    continue
                text = _row_text(values, sheet_name=sheet_name, row_index=row_offset)
                blocks.append(
                    Block(
                        block_id=block_id(document_id, 1, reading_order, "table"),
                        document_id=document_id,
                        page_number=1,
                        block_type="table",
                        text=text,
                        markdown=text,
                        reading_order=reading_order,
                        heading_path=[sheet_name],
                        metadata={
                            "source_parser": self.name,
                            "sheet_name": sheet_name,
                            "row_index": row_offset,
                            "columns": headers,
                            "row": values,
                        },
                    )
                )
                reading_order += 1

        page = Page(page_number=1, block_ids=[block.block_id for block in blocks])
        report = ParseReport(
            page_count=1,
            block_count=len(blocks),
            table_count=len(blocks),
            figure_count=0,
        )
        return ParsedDocument(
            document_id=document_id,
            source_path=str(path_obj.resolve()),
            filename=path_obj.name,
            file_type=suffix.lstrip("."),
            document_type="table_data",
            parser_used=self.name,
            parser_fallback_chain=[self.name],
            pages=[page],
            blocks=blocks,
            parse_report=report,
            metadata={"layout_source": "table_file_rows"},
        )


def _read_delimited(path: Path, delimiter: str | None) -> list[list[str]]:
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    if delimiter is not None:
        reader = csv.reader(raw.splitlines(), delimiter=delimiter)
    else:
        sample = raw[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(raw.splitlines(), dialect)
    return [[cell.strip() for cell in row] for row in reader]


def _read_xlsx(path: Path) -> dict[str, list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX parsing requires optional dependency openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[list[str]]] = {}
    for sheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        result[sheet.title] = rows
    return result


def _headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        header = value.strip() or f"column_{index}"
        if header in headers:
            header = f"{header}_{index}"
        headers.append(header)
    return headers


def _row_dict(headers: list[str], row: list[str]) -> dict[str, str]:
    padded = row + [""] * max(0, len(headers) - len(row))
    return {header: padded[index].strip() if index < len(padded) else "" for index, header in enumerate(headers)}


def _row_text(values: dict[str, Any], *, sheet_name: str, row_index: int) -> str:
    lines = [f"Sheet: {sheet_name}", f"Row: {row_index}"]
    lines.extend(f"{column}: {value}" for column, value in values.items() if str(value).strip())
    return "\n".join(lines)
